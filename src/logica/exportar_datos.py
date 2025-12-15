"""
Módulo de exportación de datos del inventario a formatos compatibles con Excel.

Soporta exportación a:
- CSV: Formato simple, abre directamente en Excel
- XLSX: Formato nativo de Excel con formato enriquecido
"""

import csv
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Tuple


def exportar_a_csv(
    datos: List[Tuple],
    columnas: List[str],
    ruta_archivo: Optional[str] = None
) -> str:
    """Exporta datos a formato CSV compatible con Excel.
    
    Args:
        datos: Lista de tuplas con los datos a exportar
        columnas: Lista con nombres de columnas
        ruta_archivo: Ruta donde guardar el archivo. Si es None, usa nombre por defecto.
    
    Returns:
        Ruta del archivo generado
    
    Example:
        >>> datos = [(1, "PC1", "Juan"), (2, "PC2", "María")]
        >>> columnas = ["DTI", "Modelo", "Usuario"]
        >>> ruta = exportar_a_csv(datos, columnas)
        >>> print(f"Archivo guardado en: {ruta}")
    """
    # Generar nombre por defecto si no se proporciona
    if not ruta_archivo:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
        ruta_archivo = output_dir / f"inventario_{timestamp}.csv"
    
    ruta_archivo = Path(ruta_archivo)
    
    # Escribir CSV con encoding UTF-8 BOM para Excel
    with open(ruta_archivo, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        
        # Escribir encabezados
        writer.writerow(columnas)
        
        # Escribir datos
        for fila in datos:
            # Convertir None a cadena vacía para mejor visualización
            fila_limpia = [str(v) if v is not None else "" for v in fila]
            writer.writerow(fila_limpia)
    
    return str(ruta_archivo)


def exportar_a_xlsx(
    datos: List[Tuple],
    columnas: List[str],
    ruta_archivo: Optional[str] = None,
    nombre_hoja: str = "Inventario"
) -> str:
    """Exporta datos a formato XLSX (Excel nativo) con formato enriquecido.
    
    Requiere el paquete openpyxl instalado.
    
    Args:
        datos: Lista de tuplas con los datos a exportar
        columnas: Lista con nombres de columnas
        ruta_archivo: Ruta donde guardar el archivo. Si es None, usa nombre por defecto.
        nombre_hoja: Nombre de la hoja de Excel
    
    Returns:
        Ruta del archivo generado
    
    Raises:
        ImportError: Si openpyxl no está instalado
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError(
            "El paquete 'openpyxl' no está instalado. "
            "Instálelo con: pip install openpyxl"
        )
    
    # Generar nombre por defecto si no se proporciona
    if not ruta_archivo:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
        ruta_archivo = output_dir / f"inventario_{timestamp}.xlsx"
    
    ruta_archivo = Path(ruta_archivo)
    
    # Crear libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir encabezados
    for col_idx, columna in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=columna)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Escribir datos
    for row_idx, fila in enumerate(datos, 2):
        for col_idx, valor in enumerate(fila, 1):
            # Convertir None a cadena vacía
            valor_limpio = valor if valor is not None else ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=valor_limpio)
            cell.border = thin_border
            
            # Centrar números y booleanos
            if isinstance(valor, (int, float, bool)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar ancho de columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max 50 caracteres
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila (encabezados)
    ws.freeze_panes = ws['A2']
    
    # Guardar archivo
    wb.save(ruta_archivo)
    
    return str(ruta_archivo)


def exportar_dispositivos_completo(
    conexion,
    formato: str = "csv",
    incluir_inactivos: bool = True
) -> str:
    """Exporta la tabla completa de dispositivos.
    
    Args:
        conexion: Conexión a la base de datos SQLite
        formato: 'csv' o 'xlsx'
        incluir_inactivos: Si True, incluye dispositivos inactivos
    
    Returns:
        Ruta del archivo generado
    """
    cursor = conexion.cursor()
    
    # Consulta SQL
    if incluir_inactivos:
        query = """
            SELECT 
                serial,
                DTI,
                user,
                MAC,
                model,
                processor,
                GPU,
                RAM,
                disk,
                CASE WHEN license_status = 1 THEN 'Sí' ELSE 'No' END as license_status,
                ip,
                CASE WHEN activo = 1 THEN 'Sí' ELSE 'No' END as activo
            FROM Dispositivos
            ORDER BY DTI, serial
        """
    else:
        query = """
            SELECT 
                serial,
                DTI,
                user,
                MAC,
                model,
                processor,
                GPU,
                RAM,
                disk,
                CASE WHEN license_status = 1 THEN 'Sí' ELSE 'No' END as license_status,
                ip,
                CASE WHEN activo = 1 THEN 'Sí' ELSE 'No' END as activo
            FROM Dispositivos
            WHERE activo = 1
            ORDER BY DTI, serial
        """
    
    cursor.execute(query)
    datos = cursor.fetchall()
    
    # Nombres de columnas en español
    columnas = [
        "Serial",
        "DTI",
        "Usuario",
        "MAC",
        "Modelo",
        "Procesador",
        "GPU",
        "RAM (GB)",
        "Disco",
        "Licencia",
        "IP",
        "Activo"
    ]
    
    # Exportar según formato
    if formato.lower() == "xlsx":
        return exportar_a_xlsx(datos, columnas, nombre_hoja="Dispositivos")
    else:
        return exportar_a_csv(datos, columnas)


def exportar_con_estado_actual(conexion, formato: str = "csv") -> str:
    """Exporta dispositivos con su estado actual (encendido/apagado).
    
    Args:
        conexion: Conexión a la base de datos SQLite
        formato: 'csv' o 'xlsx'
    
    Returns:
        Ruta del archivo generado
    """
    cursor = conexion.cursor()
    
    query = """
        SELECT 
            d.serial,
            d.DTI,
            d.user,
            d.model,
            d.processor,
            d.RAM,
            d.ip,
            CASE WHEN a.powerOn = 1 THEN 'Encendido' ELSE 'Apagado' END as estado,
            datetime(a.date, 'localtime') as ultima_verificacion
        FROM Dispositivos d
        LEFT JOIN (
            SELECT Dispositivos_serial, powerOn, date
            FROM activo
            WHERE (Dispositivos_serial, date) IN (
                SELECT Dispositivos_serial, MAX(date)
                FROM activo
                GROUP BY Dispositivos_serial
            )
        ) a ON d.serial = a.Dispositivos_serial
        WHERE d.activo = 1
        ORDER BY d.DTI, d.serial
    """
    
    cursor.execute(query)
    datos = cursor.fetchall()
    
    columnas = [
        "Serial",
        "DTI",
        "Usuario",
        "Modelo",
        "Procesador",
        "RAM (GB)",
        "IP",
        "Estado",
        "Última Verificación"
    ]
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"inventario_estado_{timestamp}"
    
    if formato.lower() == "xlsx":
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
        ruta = output_dir / f"{nombre_archivo}.xlsx"
        return exportar_a_xlsx(datos, columnas, str(ruta), nombre_hoja="Estado Actual")
    else:
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
        ruta = output_dir / f"{nombre_archivo}.csv"
        return exportar_a_csv(datos, columnas, str(ruta))
